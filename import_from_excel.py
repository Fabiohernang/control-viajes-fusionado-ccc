from datetime import datetime
from decimal import Decimal
from pathlib import Path

import openpyxl

from app import app, db, Viaje, get_config_decimal


EXCEL_PATH = Path("Control de viajes.xlsx")


def normalize_text(value):
    if value is None:
        return None
    return str(value).strip()


def to_decimal(value, default="0"):
    if value is None or value == "":
        return Decimal(default)
    return Decimal(str(value).replace(",", "."))


with app.app_context():
    if not EXCEL_PATH.exists():
        raise SystemExit(f"No encuentro el archivo: {EXCEL_PATH.resolve()}")

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    iva = get_config_decimal("iva_rate", "0.21")
    socio_rate = get_config_decimal("socio_commission_rate", "0.06")
    no_socio_rate = get_config_decimal("no_socio_commission_rate", "0.10")
    lucas_rate = get_config_decimal("lucas_commission_rate", "0.015")

    total = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [ws.cell(1, c).value for c in range(1, 18)]
        has_liquidado = "LIQUIDADO" in headers

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0] or not row[1] or not row[3]:
                continue

            fecha = row[0].date() if hasattr(row[0], "date") else datetime.strptime(str(row[0]), "%Y-%m-%d").date()

            if has_liquidado:
                liquidado = bool(row[12])
                tarifa = row[9]
                descuento = row[10]
                kg = row[11]
                observaciones = row[17] if len(row) > 17 else None
            else:
                liquidado = False
                tarifa = row[9]
                descuento = row[10]
                kg = row[11]
                observaciones = row[16] if len(row) > 16 else None

            viaje = Viaje(
                fecha=fecha,
                cliente=normalize_text(row[1]) or "",
                factura=normalize_text(row[2]),
                fletero=normalize_text(row[3]) or "",
                socio=(normalize_text(row[4] or "").upper() == "SI"),
                ctg=normalize_text(row[5]),
                origen=normalize_text(row[6]),
                destino=normalize_text(row[7]),
                kilometros=to_decimal(row[8] or 0),
                tarifa=to_decimal(tarifa or 0),
                descuento=to_decimal(descuento or 0),
                kg=to_decimal(kg or 0),
                liquidado=liquidado,
                observaciones=normalize_text(observaciones),
            )
            viaje.recalcular(iva=iva, socio_rate=socio_rate, no_socio_rate=no_socio_rate, lucas_rate=lucas_rate)

            db.session.add(viaje)
            total += 1

    db.session.commit()
    print(f"Importación terminada. Viajes cargados: {total}")
